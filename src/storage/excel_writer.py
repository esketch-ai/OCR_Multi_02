# -*- coding: utf-8 -*-
import os
import logging
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

# Column headers matching the required output format
HEADERS = [
    '차량번호',
    '업체명',
    '차대번호',
    '차명',
    '연식',
    '차량등록일',
    '차종',
    '길이(mm)',
    '너비(mm)',
    '높이(mm)',
    '총중량(kg)',
    '승차정원',
    '연료',
]


class ExcelWriter:
    """Writes OCR results to a local Excel (.xlsx) file."""

    def __init__(self, output_path):
        self.output_path = output_path
        self._ensure_workbook()

    def _ensure_workbook(self):
        """Load existing workbook or create a new one with headers."""
        if os.path.exists(self.output_path):
            self.wb = load_workbook(self.output_path)
            self.ws = self.wb.active
            logger.info(f"Loaded existing Excel file: {self.output_path}")
        else:
            os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = '자동차등록증_OCR'
            self.ws.append(HEADERS)
            self.wb.save(self.output_path)
            logger.info(f"Created new Excel file: {self.output_path}")

    def append_row(self, data):
        """
        Append a row of data to the Excel file.

        Args:
            data: list of values matching HEADERS column order
        Returns:
            True on success, False on failure
        """
        try:
            self.ws.append(data)
            logger.info(f"Row appended to {self.output_path}")
            return True
        except Exception as e:
            logger.error(f"Failed to append row to Excel: {e}")
            return False

    def close(self):
        """Close the workbook."""
        try:
            self.wb.save(self.output_path)
            self.wb.close()
        except Exception as e:
            logger.error(f"Failed to close Excel workbook: {e}")
